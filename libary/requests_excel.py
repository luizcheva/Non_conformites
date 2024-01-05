import openpyxl

DIR = (
    "\\\\straumann.com\\public\\br03\\pcoudir\\Controle de Qualidade\\"
    "01. 4400\\01. CQ Geral\\05. Banco de dados\\"
)
NAME_FILE = "Atualização Familia_p_item.xlsm"
FILE = DIR + NAME_FILE


class BuscarDescricao():
    def __init__(self) -> None:
        self.planilha = openpyxl.load_workbook(FILE).active

    def buscarItem(self, valor_pesquisar: str):
        for linha in self.planilha.iter_rows(
            min_row=2, max_row=self.planilha.max_row, min_col=1, max_col=1
        ):
            for celula in linha:
                if valor_pesquisar in str(celula.value):
                    descricaoItem = self.planilha.cell(
                        row=celula.row, column=2
                    ).value
                    return descricaoItem
        return None

    def closePlan(self):
        self.planilha.close()


if __name__ == '__main__':
    result = BuscarDescricao()
    descricao = result.buscarItem('109.1021-1')
    print(descricao)
